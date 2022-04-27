from .response import Response
from ..common.dict_functions import flatten_dict
import openpyxl
from django.http import HttpResponse
from ..common.column_letters import get_column_letters

STYLE_CENTERED_TEXT = openpyxl.styles.alignment.Alignment(horizontal="center", vertical="center")


class ExcelResponse(Response):

    def __init__(self, request, title=""):
        super().__init__(request, title)
        self.workbook = openpyxl.Workbook()

    def add_content(self, data, label=""):
        if label == "": label = self.title

        self.content[label] = data
        self.create_sheet_from_data(label, data)

    def create_sheet(self, sheet_name):
        if "Sheet" in self.workbook.sheetnames:
            sheet = self.workbook["Sheet"]
            sheet.title = sheet_name
            return sheet
        else:
            self.workbook.create_sheet(sheet_name)
            return self.workbook[sheet_name]

    def create_sheet_from_data(self, sheet_name, data, headers=None, multirow_headers=True):
        sheet = self.create_sheet(sheet_name)

        # Get fields (list) and headers (dict {field: header})
        if isinstance(headers, list):
            fields_ = headers
            headers_ = {x: x for x in headers}
        elif isinstance(headers, dict):
            fields_ = list(headers.keys())
            headers_ = headers
        else:
            fields_ = []
            for i in range(0, len(data)):
                data[i] = flatten_dict(data[i])
                fields_ += [x for x in data[i] if x not in fields_]
            headers_ = {x: x for x in fields_}

        # Get column letters
        COLUMN_LETTERS = get_column_letters(len(headers_))

        # Merge headers for multirow headers
        # This will create one row header for each nested level in the headers
        # For example, if headers are [A.a, A.b, B], then the excel should
        # look like this:
        # +-----------------+
        # |     A     |     |
        # +-----+-----+  B  |
        # |  a  |  b  |     |
        # +-----+-----+-----+
        headers_list = [headers_[x] for x in fields_]
        if multirow_headers:

            # Create a list with each row header from the headers_
            # For example, if headers_ is [A.a, A.b, B], the headers_aux
            # will be [[A, A, B], [a, b, B]]
            max_nesting_level = 0
            headers_aux = []
            for header in headers_list:
                aux = header.split(".")
                if len(aux) > max_nesting_level: max_nesting_level = len(aux)
                headers_aux.append(aux)

            for aux in headers_aux:
                while len(aux) < max_nesting_level: aux.append(aux[-1])

            headers_aux = list(map(list, zip(*headers_aux)))

            # Add headers to excel
            for row_header in headers_aux: sheet.append(row_header)

            # Merge headers horizontally
            # FIXME: Sometimes the cells are incorrectly merged when, for example, + A | B + x | x + (merges the x's)
            for r in range(0, len(headers_aux)):
                aux_c = -1
                row_len = len(headers_aux[r])

                # For each cell (c) in row
                for c in range(0, row_len):
                    # If current cell is last cell or next cell's (c + 1) value is different from aux_value
                    if c + 1 == row_len or headers_aux[r][c+1] != headers_aux[r][c]:
                        # If there is a gap greater than one cell, merge and center
                        if aux_c + 1 < c:
                            sheet.merge_cells(f'{COLUMN_LETTERS[aux_c+1]}{r+1}:{COLUMN_LETTERS[c]}{r+1}')
                        aux_c = c

            # Merge headers vertically
            for c in range(0, len(headers_aux[0])):
                aux_r = -1
                col_len = len(headers_aux)

                for r in range(0, col_len):
                    if r + 1 == col_len or headers_aux[r+1][c] != headers_aux[r][c]:
                        if aux_r + 1 < r:
                            sheet.merge_cells(f'{COLUMN_LETTERS[c]}{aux_r+2}:{COLUMN_LETTERS[c]}{r+1}')
                        aux_r = r

                    # Also center cells
                    sheet[f'{COLUMN_LETTERS[c]}{r+1}'].alignment = STYLE_CENTERED_TEXT

        else:
            # If not multirow headers, add headers to sheet normally
            sheet.append(headers_list)

        # Add data to sheet
        for record in data:
            if headers is not None: record = flatten_dict(record)
            sheet.append([record.pop(x, None) for x in fields_])

        # Prettify sheet by resizing columns
        openpyxl.worksheet.dimensions.ColumnDimension(sheet, bestFit=True)
        for column in sheet.columns:
            length = max(len(str(cell.value) if cell.value is not None else "") for cell in column)
            sheet.column_dimensions[column[-1].column_letter].width = length * 1.15

            if isinstance(column[-1].value, float):
                sheet.column_dimensions[column[-1].column_letter].number_format = "0.0"
                # You can customize the number format later on with
                # response.workbook["sheet_name"].column_dimensions("A").number_format

        return

    def make(self):
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=' + self.title + '.xlsx'
        self.workbook.save(response)
        return response
